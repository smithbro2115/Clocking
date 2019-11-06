import openpyxl
from openpyxl import styles
from LocalFileHandling import make_folder_if_it_does_not_exist, get_app_data_folder
from datetime import datetime
from Company import load_companies, Company
from utils import resource_path
from PyQt5 import QtWidgets
import os


invoice_template_path = resource_path("Invoice.xlsx")
align = styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
font = styles.Font(size=12)
try:
    default_company = load_companies(f"{get_app_data_folder('Companies')}/company_company.csv")[0]
except FileNotFoundError:
    default_company = Company()


def make_invoice_excel(user, categories, path=None):
    excel_path = make_new_excel_from_template(user, path)
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    add_user_info_to_invoice(sheet, user)
    add_company_info_to_invoice(sheet, default_company)
    for category in categories:
        add_category_to_invoice(category, sheet)
    wb.save(excel_path)


def unmerge_cells(sheet):
    for merged_cells in sheet.merged_cells.ranges:
        sheet.unmerge_cells(str(merged_cells))


def merge_cells(sheet, cells_tom_merge):
    for cell_to_merge in cells_tom_merge:
        sheet.merge_cells(str(cell_to_merge))


def verify_path(path):
    file_path, file_extension = os.path.splitext(path)
    if file_extension is not '.xlsx':
        path = file_path + ".xlsx"
    return path


def add_user_info_to_invoice(df, user):
    date = datetime.now()
    df.cell(row=4, column=2).value = date.date()
    df.cell(row=6, column=2).value = f"{user.first_name.capitalize()} {user.last_name.capitalize()}"
    df.cell(row=8, column=2).value = user.address
    df.cell(row=12, column=2).value = user.phone_number
    df.cell(row=14, column=2).value = user.email
    for column in range(1, 5):
        for row in range(4, 15):
            cell = df.cell(row=row, column=column)
            cell.alignment = align
            cell.font = font


def add_company_info_to_invoice(df, company):
    df.cell(row=1, column=6).value = f"{company.name}\n{company.motto}"
    df.cell(row=6, column=7).value = company.name
    df.cell(row=8, column=7).value = company.address
    df.cell(row=12, column=7).value = company.phone_number
    df.cell(row=14, column=7).value = company.email
    for column in range(6, 10):
        for row in range(1, 15):
            cell = df.cell(row=row, column=column)
            cell.alignment = align
            cell.font = font


def add_category_to_invoice(category, df):
    for i in range(18, 28):
        if not df.cell(row=i, column=1).value:
            date = datetime.now()
            df.cell(row=i, column=1).value = date.date()
            df.cell(row=i, column=2).value = round(float(category.clock.total_monthly_time.total_seconds())/3600, 4)
            df.cell(row=i, column=3).value = category.category_number
            df.cell(row=i, column=4).value = category.description
            df.cell(row=i, column=8).value = category.wage
            return True
    return False


def make_new_excel_from_template(user, path=None):
    import shutil
    if path:
        new_path = path
    else:
        new_dir = f"{user.directory}/Invoices"
        make_folder_if_it_does_not_exist(user.directory, 'Invoices')
        new_path = f"{new_dir}/{get_file_invoice_name(user)}"
    new_path = verify_path(new_path)
    shutil.copy(invoice_template_path, new_path)
    return new_path


def get_file_invoice_name(user):
    date = datetime.now()
    return f"{user.first_name}_{user.last_name}_Invoice" \
        f"_{date.strftime('%B')}_{date.day}_{date.year}.xlsx"


def get_invoice_folder_name():
    date = datetime.now()
    return f"Invoices_{date.strftime('%B')}_{date.year}"


class GetFileLocationDialog(QtWidgets.QFileDialog):
    def __init__(self, default_name, caption, default_location=None):
        super(GetFileLocationDialog, self).__init__()
        self.caption = caption
        self.default_location = default_location if default_location else '/'
        self.default_name = default_name

    def get_save_path(self):
        result = self.getSaveFileName(directory=f'{self.default_location}{self.default_name}', caption=self.caption)[0]
        return result
